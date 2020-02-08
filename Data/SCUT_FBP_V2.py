import numpy as np
import cv2
import os
import codecs
import re
import struct
import matplotlib.pyplot as plt
import openpyxl as xls
import pickle
from collections import OrderedDict


def main():
    root_dir = 'D:/ProgramData/SCUT-FBP5500_v2'

    racial = ['AF', 'CF', 'AM', 'CM']
    num_images = [2000, 750, 2000, 750]
    csv_sheet = ['Asian_Female', 'Caucasian_Female', 'Asian_Male', 'Caucasian_Male']
    
    label = []
    csv_file = xls.load_workbook( os.path.join(root_dir, 'All_Ratings.xlsx') )
    for s, sheet in enumerate(csv_sheet):
        print(s, sheet)
        label.append( load_label(csv_file, sheet, 60, num_images[s]) )
    
    label = np.concatenate(label, axis=0)

    annotation = []
    l = 0
    for r, race in enumerate(racial):
        for i in range(num_images[r]):
            this_anno = {
                'path': 'Images/' + race + str(i+1) + '.jpg',
                'all_ratings': label[l, :],
                'landmarks': load_landmarkds( os.path.join(root_dir, 'facial landmark', race+str(i+1)+'.pts') )
            }
            annotation.append(this_anno)
            l += 1
            print(race, i+1)
    
    print(len(annotation))
    print(annotation[0]['path'])
    print(annotation[0]['landmarks'].shape)
    print(annotation[0]['all_ratings'].shape)

    pk_file = os.path.join(root_dir, 'Annotation.pkl')
    with open(pk_file, 'wb') as outFile:
        pickle.dump(annotation, outFile)
            


def test():
    root_dir = 'D:/ProgramData/SCUT-FBP5500_v2'
    
    pkl_file = os.path.join(root_dir, 'Annotation.pkl')
    with open(pkl_file, 'rb') as f:
        annotation = pickle.load(f)

    data = annotation[5100]

    img = cv2.imread(os.path.join(root_dir, data['path']))
    landmarks = data['landmarks'].astype(np.int32)
    ratings = data['all_ratings']
    print(data['path'])

    for i in range(landmarks.shape[0]):
        cv2.circle(img, (landmarks[i, 0], landmarks[i, 1]), 2, (255,0,0), 2) 

    print(ratings)

    cv2.imshow('1', img)
    cv2.waitKey(0)


def train_test_divide(total, train, test):
    index = np.arange(total)
    np.random.shuffle(index)
    
    data_index = OrderedDict()
    data_index['train'] = index[:train]
    data_index['test'] = index[train:train+test]

    # Save in pkl
    pk_file = 'D:/ThesisData/SCUT-FBP5500_v2/data_index_1800_200.pkl'
    with open(pk_file, 'wb') as outFile:
        pickle.dump(data_index, outFile)




def load_label(wb, sheet, raters, images):
        s1 = wb[sheet]

        label = np.zeros((images, raters))
        rows = raters * images + 1

        for i in range(2, rows + 1):
            num_rater = s1.cell(row=i, column=1).value
            num_image = s1.cell(row=i, column=2).value
            num_image = int(num_image[2:-4])
            score = s1.cell(row=i, column=3).value
            label[num_image - 1, num_rater - 1] = score

        return label



def load_landmarkds(path):
    data = open(path, 'rb').read()
    points = struct.unpack('i172f', data)
    points = points[1:]
    num_points = int(len(points)/2)
    arr = np.zeros([num_points, 2])
    for p in range(num_points):
        arr[p, 0] = points[2*p]
        arr[p, 1] = points[2*p+1]

    return arr



if __name__ == '__main__':
    train_test_divide(2000, 1800, 200)



'''def onclick(event):
    x = np.int32(event.xdata)
    y = np.int32(event.ydata)
    landmarks.append(x)
    landmarks.append(y)
    if len(landmarks) == 173:
        points = struct.pack('i172f', *landmarks)
        newFile = open('CM152.pts', 'wb')
        newFile.write(points)
    else:
        print((len(landmarks) - 1) / 2)
        cv2.circle(img, (x, y), 2, (255,0,0), 2) 
        ax.imshow(img)
        plt.draw()



landmarks = []
fig, ax = plt.subplots()
landmarks.append(np.int32(68))
root_dir = 'D:/ProgramData/SCUT-FBP5500_v2/Images/CM152.jpg'
img = cv2.imread(root_dir)

fig.canvas.mpl_connect('button_press_event', onclick)
ax.imshow(img)
plt.show()'''


        
