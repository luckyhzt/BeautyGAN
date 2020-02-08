import os
import numpy as np
import openpyxl as xls


def load_label(file, sheet, raters, images):
        wb = xls.load_workbook(file)
        s1 = wb[sheet]

        label = np.zeros((images, raters))
        rows = s1.max_row

        for i in range(2, rows + 1):
            num_rater = s1.cell(row=i, column=1).value
            num_image = s1.cell(row=i, column=2).value
            score = s1.cell(row=i, column=3).value
            label[num_image - 1, num_rater - 1] = score

        return label

def average_score(label):
    avg = np.sum(label, axis=1)
    num = np.count_nonzero(label, axis=1)
    avg = avg / num

    return avg

def label_distribution(label):
    label = label.astype(int)
    images = label.shape[0]
    dist = np.zeros((images, 5))

    for i in range(0, images):
        occur = np.bincount(label[i, :].reshape((-1)), minlength=6)
        occur = occur[1:6]
        dist[i, :] = occur / np.sum(occur)

    return dist

def label_level(label, level_gaps):
    level = np.floor( (label - 1) / level_gaps)
    level = one_hot(level, num_class=int(4.0/level_gaps))

    return level


def one_hot(label, num_class):
    label = label.astype(int)
    n = label.shape[0]
    onehot = np.zeros((n, num_class), np.float32)
    
    for i in range(n):
        onehot[i, label[i]] = 1.0

    return onehot


if __name__ == '__main__':
    label = load_label('./Data/Labeled_data/Label/Ratings of all raters.xlsx', 'Sheet1', 75, 500)
    label_avg = average_score(label)
    label_avg = np.expand_dims(label_avg, axis=1)
    label_dist = label_distribution(label)
    label_lvl = label_level(label_avg, level_gaps=1.0)

    print(np.sum(label_lvl[400:500, :], axis=0))

    #np.save('./Data/Labeled_data/Label/label_avg.npy', label_avg)
    #np.save('./Data/Labeled_data/Label/label_dist.npy', label_dist)


    